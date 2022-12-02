from array import array
from ast import Break
from operator import is_not
from typing import List
import pandas as pd
import os
import time
import sys
import csv
import re
import xlsxwriter
import pathlib
from openpyxl import load_workbook
import pandas.io.formats.excel

def LeerArchivo(Ruta):
    Path='\\'.join(Ruta.split('\\')[:-1])
    ListaTablas=[]
    ejemplo_dir = Path
    directorio = pathlib.Path(ejemplo_dir)
    dfFinal = pd.DataFrame()
    dfFinal['CFDI_UUID_FACTURA_TPGO']=None
    dfFinal['CFDI_UUID_COMPLEMENTO_PAGO_TPGO']=None
    dfFinal['MONTO_APLICADO_TPGO']=None
    dfFinal['DIFERENCIA_TPGO']=None
    for fichero in directorio.iterdir():
        
        if "PAGOS" in str(fichero) and "$" not in str(fichero):
            Fichero2 = str(fichero)
            Fichero2=str(fichero).split('_')
            print('Este es el numero del archivo de pagos ',Fichero2[1])
    print('Comienza Lectura Archivos UUID')        
    for fichero in directorio.iterdir():
        if "UUIDS" in str(fichero.name) and "$" not in str(fichero) :
            Fichero3 = str(fichero)
            Fichero3=str(fichero).split('_')
            print('Este es el numero de los archivos UUID',Fichero3[1])
            if (Fichero2[1]==Fichero3[1]):
               print('este es el fichero uuid Correspondiente ',Fichero3[1])
               df_UuidTemp=pd.read_excel(fichero,sheet_name='TOTALPLAY')
               df_UuidTemp.drop(['EMPRESA','ACCOUNT_NO','MONTO_PAGO','MONTO_COMPLEMENTO','APLICACION','BILL_NO','FECHA_FACTURA'], axis = 1, inplace = True)
               ListaTablas.append(df_UuidTemp)
    print('Termina Lectura Archivos UUID')
    print("----------------------------------")
    print("Comienza lectura PAGOS")
    for fichero in directorio.iterdir():
       
        if "PAGOS" in str(fichero) and "$" not in str(fichero):
            df_PagosActual=pd.read_excel(fichero,sheet_name='MX')
            df_PagosActual.drop(['NUM','EMPRESA','ACCOUNT_NO','COMPANIA','NOMBRE','RFC','MONTO_PAGO','FECHA_LIQUIDADO','TPI_MONTO','TPI_IEPS','TPI_IVA','TPT_MONTO','TPT_IEPS','TPT_IVA','HS_MONTO','HS_IEPS','HS_IVA','TB_MONTO','TB_IEPS','TB_IVA','TERCEROS','PENDIENTE','CANAL','CUENTA_BANCO','MONEDA','COUNTRY','UBICACION','Código Postal'],axis = 1, inplace = True)
            for i in range(0,df_PagosActual.shape[0]):
                Item=df_PagosActual['ITEM_NO'][i]
                print("Item: ",Item," - ",i," de ",df_PagosActual.shape[0])
                Bandera = False
                for j in range(0,len(ListaTablas)):
                    if(Bandera == False):
                        Tabla = ListaTablas[j].loc[ListaTablas[j]['ITEM_NO'] == Item,['MONTO_APLICADO','CFDI_UUID_FACTURA','CFDI_UUID_COMPLEMENTO_PAGO']]
                       
                        if Tabla.shape[0] == 0 and j+1 == len(ListaTablas):
                            dfFinal.loc[i]=["UUID NO ENCONTRADO","UUID NO ENCONTRADO","0.0","0.0"]
                           
                            Bandera = True
                        
                        if Tabla.shape[0] >= 1:
                          
                            listaImporte = Tabla.MONTO_APLICADO
                            listaFacturas = Tabla.CFDI_UUID_FACTURA
                            listaComplemento = Tabla.CFDI_UUID_COMPLEMENTO_PAGO
                            CadenaFacturas=''
                            for item in listaFacturas:
                                CadenaFacturas = CadenaFacturas+str(item)+'/'
                            Sumatoria=float(df_PagosActual['GO_MONTO'][i])+float(df_PagosActual['GO_IEPS'][i])+float(df_PagosActual['GO_IVA'][i])
                            diferencia = listaImporte.sum()-float(Sumatoria)
                       
                            Factura=listaComplemento.iloc[0]
                            CFDI=Factura
                            dfFinal.loc[i]=[CadenaFacturas,CFDI,float(listaImporte.sum()),float(diferencia)]
                            Bandera = True
                          
                      
                    else:
                        
                        Break    
            print("COMIENZA ESCRITURA")
            book = load_workbook(fichero)
            writer = pd.ExcelWriter(fichero, engine='openpyxl')
            writer.book = book
            pandas.io.formats.excel.header_style = None
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            dfFinal.to_excel(writer, 'MX', startcol = 42,  index = False)
            writer.save()
            dfFinal.reset_index().to_csv('fichero.csv',header=True,index=False)         
            print('------------------------------------')
            
    dfFinal = pd.DataFrame()
    dfFinal['CFDI_UUID_FACTURA_TPGO']=None
    dfFinal['CFDI_UUID_COMPLEMENTO_PAGO_TPGO']=None
    dfFinal['MONTO_APLICADO_TPGO']=None
    dfFinal['DIFERENCIA_TPGO']=None
    
    print("Comienza lectura PAGOS_NORTE")
    
    for fichero in directorio.iterdir():
       
        if "PAGOS" in str(fichero) and "$" not in str(fichero):
            df_PagosActual=pd.read_excel(fichero,sheet_name='MX_NORTE')
            df_PagosActual.drop(['NUM','EMPRESA','ACCOUNT_NO','COMPANIA','NOMBRE','RFC','MONTO_PAGO','FECHA_LIQUIDADO','TPI_MONTO','TPI_IEPS','TPI_IVA','TPT_MONTO','TPT_IEPS','TPT_IVA','HS_MONTO','HS_IEPS','HS_IVA','TB_MONTO','TB_IEPS','TB_IVA','TERCEROS','PENDIENTE','CANAL','CUENTA_BANCO','MONEDA','COUNTRY','UBICACION','Código Postal'],axis = 1, inplace = True)
            for i in range(0,df_PagosActual.shape[0]):
                Item=df_PagosActual['ITEM_NO'][i]
                print("Item: ",Item," - ",i," de ",df_PagosActual.shape[0])
                Bandera = False
                for j in range(0,len(ListaTablas)):
                    if(Bandera == False):
                        Tabla = ListaTablas[j].loc[ListaTablas[j]['ITEM_NO'] == Item,['MONTO_APLICADO','CFDI_UUID_FACTURA','CFDI_UUID_COMPLEMENTO_PAGO']]
                       
                        if Tabla.shape[0] == 0 and j+1 == len(ListaTablas):
                            dfFinal.loc[i]=["UUID NO ENCONTRADO","UUID NO ENCONTRADO","0.0","0.0"]
                           
                            Bandera = True
                        
                        if Tabla.shape[0] >= 1:
                          
                            listaImporte = Tabla.MONTO_APLICADO
                            listaFacturas = Tabla.CFDI_UUID_FACTURA
                            listaComplemento = Tabla.CFDI_UUID_COMPLEMENTO_PAGO
                            CadenaFacturas=''
                            for item in listaFacturas:
                                CadenaFacturas = CadenaFacturas+str(item)+'/'
                            Sumatoria=float(df_PagosActual['GO_MONTO'][i])+float(df_PagosActual['GO_IEPS'][i])+float(df_PagosActual['GO_IVA'][i])
                            diferencia = listaImporte.sum()-float(Sumatoria)
                       
                            Factura=listaComplemento.iloc[0]
                            CFDI=Factura
                            dfFinal.loc[i]=[CadenaFacturas,CFDI,float(listaImporte.sum()),float(diferencia)]
                            Bandera = True
                          
                      
                    else:
                        
                        Break    
            print("COMIENZA ESCRITURA")
            book = load_workbook(fichero)
            writer = pd.ExcelWriter(fichero, engine='openpyxl')
            writer.book = book
            pandas.io.formats.excel.header_style = None
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            dfFinal.to_excel(writer, 'MX_NORTE', startcol = 42,  index = False)
            writer.save()
            dfFinal.reset_index().to_csv('fichero.csv',header=True,index=False)         
            print('------------------------------------')
    f=open(Path+'\\salida.abc','w')
    f.close()
                
if __name__ == "__main__":
    Tstart = time.time()
    argumentList = sys.argv
    if len(argumentList) == 2:
        LeerArchivo(argumentList[1])
    else:
        print('No se cumple con el número de argumentos (2)')
    Tdone = time.time()
    elapsed = Tdone - Tstart
    print('Tiempo de ejecución: '+str(elapsed))
